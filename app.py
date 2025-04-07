import streamlit as st
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import os

st.set_page_config(page_title="Checklist Fiscalização", page_icon="✅")
st.title("📋 Checklist de Fiscalização")

arquivo_excel = "checklists.xlsx"

# Perguntas do checklist
perguntas = [
    "Qual a empresa?",
    "Qual a data?",
    "Qual atividade?",
    "Qual Local?",
    "Qual OM?",
    "Isolamento e APR ok?",
    "Epi's ok?",
    "N° de funcionários OK?",
    "Atividades da OM foram executadas?",
    "Recursos Samarco disponíveis?",
    "Atividade tem início conforme programado?"
]

# Inicializa a lista de checklists na sessão
if "respostas_todas" not in st.session_state:
    st.session_state.respostas_todas = []

# Formulário para preencher o checklist
with st.form("checklist_form"):
    st.subheader("Preencha o checklist:")
    respostas = [st.text_input(pergunta) for pergunta in perguntas]
    enviado = st.form_submit_button("Salvar")

def formatar_planilha(ws):
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 50

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for col in ["A", "B"]:
        ws[f"{col}1"].font = header_font
        ws[f"{col}1"].fill = header_fill
        ws[f"{col}1"].alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border

# Quando o usuário clica em salvar
if enviado:
    st.session_state.respostas_todas.append(respostas)
    st.success("Checklist salvo na memória! Preencha outro ou exporte todos ao final.")

# Botão para exportar todos os checklists para Excel
if st.session_state.respostas_todas:
    if st.button("📥 Baixar todos os checklists (Excel)"):
        wb = Workbook()
        del wb["Sheet"]  # remove aba padrão

        for i, respostas in enumerate(st.session_state.respostas_todas):
            nome_aba = f"{respostas[0]}_{respostas[4]}" if respostas[0] and respostas[4] else f"Checklist_{i+1}"
            if nome_aba in wb.sheetnames:
                contador = 1
                while f"{nome_aba}_{contador}" in wb.sheetnames:
                    contador += 1
                nome_aba = f"{nome_aba}_{contador}"

            ws = wb.create_sheet(title=nome_aba)
            ws.append(["Pergunta", "Resposta"])
            for pergunta, resposta in zip(perguntas, respostas):
                ws.append([pergunta, resposta])
            formatar_planilha(ws)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        st.download_button(
            label="📥 Clique aqui para baixar",
            data=buffer,
            file_name="checklists.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
